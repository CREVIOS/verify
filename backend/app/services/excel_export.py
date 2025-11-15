"""
Excel Export Service for Verification Results
Exports verification data with all details: source name, sentence, context, citations, AI reasoning
"""

from typing import List
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from app.schemas.sentences import VerifiedSentenceResponse


class ExcelExportService:
    """Service for exporting verification results to Excel"""

    @staticmethod
    def export_verification_results(
        sentences: List[VerifiedSentenceResponse],
        project_name: str
    ) -> BytesIO:
        """
        Export verification results to Excel with comprehensive details

        Args:
            sentences: List of verified sentences with citations
            project_name: Name of the project

        Returns:
            BytesIO: Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Verification Results"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        validated_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        uncertain_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        incorrect_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Define headers
        headers = [
            "Status",
            "Confidence Score",
            "Sentence / Claim",
            "Page Number",
            "Context Before",
            "Context After",
            "AI Reasoning",
            "Number of Citations",
            "Citation Sources",
            "Citation Pages",
            "Citation Excerpts",
            "Citation Similarity Scores",
            "Verification Date",
            "Model Used"
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Set column widths
        column_widths = {
            1: 15,   # Status
            2: 15,   # Confidence
            3: 60,   # Sentence
            4: 12,   # Page Number
            5: 40,   # Context Before
            6: 40,   # Context After
            7: 60,   # AI Reasoning
            8: 15,   # Number of Citations
            9: 40,   # Citation Sources
            10: 15,  # Citation Pages
            11: 60,  # Citation Excerpts
            12: 20,  # Similarity Scores
            13: 18,  # Verification Date
            14: 30   # Model Used
        }

        for col_num, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Write data rows
        for row_num, sentence in enumerate(sentences, 2):
            # Status
            status_cell = ws.cell(row=row_num, column=1)
            status_cell.value = sentence.status.value.upper()
            status_cell.border = border
            status_cell.alignment = Alignment(horizontal="center", vertical="top")

            # Apply color based on status
            if sentence.status.value == "validated":
                status_cell.fill = validated_fill
            elif sentence.status.value == "uncertain":
                status_cell.fill = uncertain_fill
            elif sentence.status.value == "incorrect":
                status_cell.fill = incorrect_fill

            # Confidence Score
            confidence_cell = ws.cell(row=row_num, column=2)
            confidence_cell.value = f"{sentence.confidence_score * 100:.1f}%"
            confidence_cell.border = border
            confidence_cell.alignment = Alignment(horizontal="center", vertical="top")

            # Sentence / Claim
            sentence_cell = ws.cell(row=row_num, column=3)
            sentence_cell.value = sentence.content
            sentence_cell.border = border
            sentence_cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Page Number
            page_cell = ws.cell(row=row_num, column=4)
            page_cell.value = sentence.page_number
            page_cell.border = border
            page_cell.alignment = Alignment(horizontal="center", vertical="top")

            # Context Before
            context_before_cell = ws.cell(row=row_num, column=5)
            context_before_cell.value = sentence.context_before or "N/A"
            context_before_cell.border = border
            context_before_cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Context After
            context_after_cell = ws.cell(row=row_num, column=6)
            context_after_cell.value = sentence.context_after or "N/A"
            context_after_cell.border = border
            context_after_cell.alignment = Alignment(wrap_text=True, vertical="top")

            # AI Reasoning
            reasoning_cell = ws.cell(row=row_num, column=7)
            reasoning_cell.value = sentence.ai_reasoning or "N/A"
            reasoning_cell.border = border
            reasoning_cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Number of Citations
            citations_count_cell = ws.cell(row=row_num, column=8)
            citations_count_cell.value = len(sentence.citations)
            citations_count_cell.border = border
            citations_count_cell.alignment = Alignment(horizontal="center", vertical="top")

            # Extract citation details
            citation_sources = []
            citation_pages = []
            citation_excerpts = []
            citation_scores = []

            for citation in sentence.citations:
                # Get document name from metadata if available
                doc_name = "Supporting Document"
                if citation.metadata and isinstance(citation.metadata, dict):
                    doc_name = citation.metadata.get("document_name", doc_name)

                citation_sources.append(doc_name)
                citation_pages.append(str(citation.page_number))
                citation_excerpts.append(citation.content[:200] + "..." if len(citation.content) > 200 else citation.content)
                citation_scores.append(f"{citation.similarity_score * 100:.1f}%")

            # Citation Sources
            sources_cell = ws.cell(row=row_num, column=9)
            sources_cell.value = "\n".join(citation_sources) if citation_sources else "N/A"
            sources_cell.border = border
            sources_cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Citation Pages
            pages_cell = ws.cell(row=row_num, column=10)
            pages_cell.value = "\n".join(citation_pages) if citation_pages else "N/A"
            pages_cell.border = border
            pages_cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Citation Excerpts
            excerpts_cell = ws.cell(row=row_num, column=11)
            excerpts_cell.value = "\n\n---\n\n".join(citation_excerpts) if citation_excerpts else "N/A"
            excerpts_cell.border = border
            excerpts_cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Citation Similarity Scores
            scores_cell = ws.cell(row=row_num, column=12)
            scores_cell.value = "\n".join(citation_scores) if citation_scores else "N/A"
            scores_cell.border = border
            scores_cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Verification Date
            date_cell = ws.cell(row=row_num, column=13)
            date_cell.value = sentence.created_at.strftime("%Y-%m-%d %H:%M") if sentence.created_at else "N/A"
            date_cell.border = border
            date_cell.alignment = Alignment(horizontal="center", vertical="top")

            # Model Used
            model_cell = ws.cell(row=row_num, column=14)
            model_used = "GPT-4.1 + Gemini 2.5 Pro"
            if sentence.metadata and isinstance(sentence.metadata, dict):
                model_used = sentence.metadata.get("model_used", model_used)
            model_cell.value = model_used
            model_cell.border = border
            model_cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Freeze the header row
        ws.freeze_panes = "A2"

        # Add metadata sheet
        metadata_ws = wb.create_sheet(title="Metadata")
        metadata_ws.cell(row=1, column=1).value = "Project Name"
        metadata_ws.cell(row=1, column=2).value = project_name
        metadata_ws.cell(row=2, column=1).value = "Export Date"
        metadata_ws.cell(row=2, column=2).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        metadata_ws.cell(row=3, column=1).value = "Total Sentences"
        metadata_ws.cell(row=3, column=2).value = len(sentences)
        metadata_ws.cell(row=4, column=1).value = "Validated"
        metadata_ws.cell(row=4, column=2).value = sum(1 for s in sentences if s.status.value == "validated")
        metadata_ws.cell(row=5, column=1).value = "Uncertain"
        metadata_ws.cell(row=5, column=2).value = sum(1 for s in sentences if s.status.value == "uncertain")
        metadata_ws.cell(row=6, column=1).value = "Incorrect"
        metadata_ws.cell(row=6, column=2).value = sum(1 for s in sentences if s.status.value == "incorrect")
        metadata_ws.cell(row=7, column=1).value = "AI Models Used"
        metadata_ws.cell(row=7, column=2).value = "GPT-4.1 (Primary) + Gemini 2.5 Pro (Validation)"

        # Style metadata sheet
        for row in range(1, 8):
            metadata_ws.cell(row=row, column=1).font = Font(bold=True)
            metadata_ws.cell(row=row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        metadata_ws.column_dimensions['A'].width = 25
        metadata_ws.column_dimensions['B'].width = 50

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file
